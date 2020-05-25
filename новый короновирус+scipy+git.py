import requests
import string
import re
import datetime
import bs4
import openpyxl
from pathlib import Path
import os
import numpy
import matplotlib.pyplot as plt
import time
import scipy.optimize
#import scipy
from IPython.display import display, Image
import vk_api
import matplotlib.ticker as mticker

#получаю инфу о кейсах с сайта минздрава
CasesTotalRF=0
while CasesTotalRF==0:
    minzdrav = requests.get('https://coronavirus-monitor.info/country/russia/')
    segodnya = datetime.datetime.now()
    segodnyastr = str(segodnya)
    Minzdrav = bs4.BeautifulSoup(minzdrav.text, 'html.parser')
    #casesinRFRegex = Minzdrav.select('div.container:nth-child(2) > p:nth-child(6) > b:nth-child(2)')
    DiedinRFRegex=Minzdrav.select('html body section#content div.container.content p b')
    #print(casesinRFRegex)
    #print(DiedinRFRegex)
    DiedinRF= DiedinRFRegex[4]
    CasesinRF = DiedinRFRegex[2]
    #print(CasesinRF)
    #CasesinRF = str(CasesinRF[0])
    #print(CasesinRF)
    CasesinRF = ''.join(str(CasesinRF).split())
    #print(DiedinRF)
    #CasesinRF = re.sub(r'\s+', '', CasesinRF, flags=re.UNICODE)
    #print(CasesinRF)
    CasesNumberRF = re.compile(r'(\d+)')
    NumberRF = CasesNumberRF.search(str(CasesinRF))
    DiedNumberRF = re.compile(r'(\d+)')
    DiedNumberinRF = DiedNumberRF.search(str(DiedinRF))
    #print(DiedNumberinRF)
    #print(NumberRF)
    CasesTotalRF = int(NumberRF.group(1))
    DiedTotalRF = int(DiedNumberinRF.group(1))
    #print(CasesTotalRF)
    #print(DiedTotalRF)
datainfy=str(segodnyastr[:10])
timeinfy=str(segodnyastr[11:16])
#print('Current date '+datainfy)
#print('Current time '+timeinfy)
#print('Current total cases = '+ str(CasesTotalRF))
#перехожу в папку с эксель файлом
def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)
path_virus=Path(find('virus.xlsx','C:\\Users'))
os.chdir(path_virus.parent)
wb = openpyxl.load_workbook('virus.xlsx')
sheet = wb['Лист1']
lastRow=str(sheet.max_row)
lastColumn=str(sheet.max_column)
LastDate=sheet['A'+lastRow].value
LastCases=sheet['B'+lastRow].value
rownumber=2
IDate=sheet['A'+str(rownumber)].value
#ищу последнюю строку которая заполнена
if any(c.isalpha() for c in str(IDate))==False:
    while IDate<=segodnya:
        
      
        rownumber=rownumber+1
        IDate=sheet['A'+str(rownumber)].value        
else:
    rownumber=rownumber+1
    IDate=sheet['A'+str(rownumber)].value
    while IDate<=segodnya:
       
      
        rownumber=rownumber+1
        IDate=sheet['A'+str(rownumber)].value     
rownumber=rownumber-1        

#вставляю в ячейку сегодня данные за сегодня
sheet.cell(row=rownumber, column=3).value = CasesTotalRF
sheet.cell(row=rownumber, column=5).value = DiedTotalRF
#надо аппроскимировать теперь экспонентой
X=[]
Y=[]
W=[]
for r in range(2, rownumber+1):
        X.append(sheet.cell(row = r, column = 2).value)
        Y.append(sheet.cell(row = r, column = 3).value)
        W.append(sheet.cell(row = r, column = 5).value)
#print('x=')
#print(X)
#print('y=')
#print(Y)
#print('w=')
#print(W)
X = numpy.array(X)
Y = numpy.array(Y)
W = numpy.array(W)
#print('w=')
#print(W)
#print('x=')
#print(X)
#print('y=')
#print(Y)
#Z1=numpy.polyfit(X, numpy.log(Y), 1, w=numpy.sqrt(Y))
def f1(x, a, b):
    y=  a*numpy.exp(b*x)
    return y
Z1=scipy.optimize.curve_fit(f1, X[17:], Y[17:],method='lm')
B1=Z1[0][1]
A1=Z1[0][0]
sheet.cell(row=2, column=8).value = A1
sheet.cell(row=3, column=8).value = B1
print('A='+str(A1))
print('B='+str(B1))
Boundlogistic1=((0,0),(100,1))
Z11=scipy.optimize.curve_fit(f1, X, W,method='trf',bounds=Boundlogistic1)
B11=Z11[0][1]
A11=Z11[0][0]



def f2(x, a, b, c,d):
    y= a / (1 + numpy.exp(b * (x + c )))+d
    return y

Boundlogistic=((0,-10,-100,-10),(1000,0,100,10))

Z21=scipy.optimize.curve_fit(f2, X, W,method='trf',bounds=Boundlogistic)
A21=Z21[0][0]
B21=Z21[0][1]
C21=Z21[0][2]
D21=Z21[0][3]
print('A21='+str(A21))
print('B21='+str(B21))
print('C21='+str(C21))
print('D21='+str(D21))

Boundlogistic=((50,-10,-100,-10000),(3000000,0,100,500))



#24437,-0,218,-25,5,6
#[20000, 100000], [ -10, 10],[ -100, 100], [ -10, 100]
#Boundlogistic=((3000,-100,-100,-10),(3000000,0,100,500))
Z2=scipy.optimize.curve_fit(f2, X, Y,method='trf',bounds=Boundlogistic)

#print(Z2[0])
#print(Z2[0][0])
#L/(1+numpy.exp(-k*(X-x0)))
# вывожу параметры для расчета в экселе
A2=Z2[0][0]
B2=Z2[0][1]
C2=Z2[0][2]
D2=Z2[0][3]
print('A='+str(A2))
print('B='+str(B2))
print('C='+str(C2))
print('D='+str(D2))
sheet.cell(row=2, column=11).value = A2
sheet.cell(row=3, column=11).value = B2
sheet.cell(row=4, column=11).value = C2
sheet.cell(row=5, column=11).value = D2
sheet.cell(row=rownumber, column=6).value = int(A2)
# фигачим тип график
x = numpy.linspace(0, rownumber, 100)
x3 = numpy.linspace(0,rownumber+10, 100)
Perday=(numpy.exp(B1)-1)*100
fig, ax = plt.subplots(2,2,figsize=(20,12),tight_layout=True,sharex='col')

ax[0][0].xaxis.set_major_locator(mticker.MultipleLocator(2))
#plt.gca().yaxis.set_major_locator(mticker.MultipleLocator(2))
#ax[0][0].get_shared_x_axes().join(ax[0][1], ax[1][0])

ax[0][0].plot(x,  A1*numpy.exp(B1*x),label='Аппроксимация-экспонента',linestyle='dashed',color='xkcd:peach')
ax[0][0].plot(x, f2(x, A2, B2, C2,D2),label='Аппроксимация-логистическая функция',linestyle='dashed',color='xkcd:olive drab')

ax[0][0].set_ylabel('Число случаев заражения')
ax[0][0].set_title('Динамика заболевания')
ax[0][0].scatter(X, Y,label='Данные Минздрава',color='xkcd:cadet blue')
ax[0][0].legend()
ax[0][0].grid(True)

ax[1][0].set_yscale('log')

ax[1][0].plot(x, 100*numpy.exp(numpy.log(1.24)*x) ,label='22% в день',linestyle='dashed',color='xkcd:red')
ax[1][0].plot(x, 280*numpy.exp(numpy.log(1.16)*x) ,label='16% в день',linestyle='dashed',color='xkcd:orange')
#ax[1][0].plot(x, A1*numpy.exp(B1*x) ,label='Аппроксимация-экспонента',linestyle='dashed',color='xkcd:peach')
ax[1][0].plot(x,f2(x, A2, B2, C2,D2) ,label='Аппроксимация-логистическая функция',linestyle='dashed',color='xkcd:olive drab')
ax[1][0].set_ylabel('Число случаев заражения log10')
ax[1][0].scatter(X, Y,label='Данные Минздрава',color='xkcd:cadet blue')
ax[1][0].legend()
ax[1][0].grid(True)
ax[1][0].set_xlabel('Число дней с 16 марта ')
ax[1][0].axis('on')
#print(W)
#ax[0][1].plot(X, W, label='Данные Минздрава',color='xkcd:cadet blue')
ax[0][1].set_ylabel('Всего умерло')
ax[0][1].plot(x+3, A11*numpy.exp(B11*(x+3)) ,label='Аппроксимация-экспонента',linestyle='dashed',color='xkcd:peach')
ax[0][1].plot(x+3, f2(x+3, A21, B21, C21,D21),label='Аппроксимация-логистическая функция',linestyle='dashed',color='xkcd:olive drab')
ax[0][1].scatter(X, W,label='Данные Минздрава',color='xkcd:cadet blue')

ax[0][1].grid(True)
ax[0][1].set_xlabel('Число дней с 16 марта ')
ax[0][1].axis('on')

ax[0][1].legend()



ax[1][1].plot(x3, A1*numpy.exp(B1*x3) ,label='Аппроксимация-экспонента',linestyle='dashed',color='xkcd:peach')
ax[1][1].plot(x3,f2(x3, A2, B2, C2,D2) ,label='Аппроксимация-логистическая функция',linestyle='dashed',color='xkcd:olive drab')
ax[1][1].scatter(X, Y,label='Данные Минздрава',color='xkcd:cadet blue')
ax[1][1].grid(True)
ax[1][1].set_ylabel('Число случаев заражения')
ax[1][1].legend()



wb.save('virus.xlsx')

img_name=str('график от '+segodnyastr[0:13]+'.png')
#print(img_name)
plt.savefig(img_name,transparent=True,format='png')

#plt.show(grafik)
#time.sleep(5)
#plt.close(grafik)

# теперь коннект к вк нужен
#код берем для бота от группы
text_file = open("key vk.txt", "r")
key_vk = str(text_file.readlines()[0])
# сообщения людям и беседам
#получаем картинки сервер


ID=[id]
print('id')
otvet=input()
target_id=0
while True:
     try:
        otvet=int(otvet)
        if otvet==1:
            target_id = my_id
            #print(target_id)
        if otvet == 2:
            target_id = chat_tf_02m_19
        if otvet == 3:
            target_id = chat_memy_tes
        if otvet == 4:
            target_id = Igor_id
        if otvet == 5:
            target_id = sahsa_id
        if otvet == 6:
            target_id = Lera_id
     except ValueError:
         print('цифры надо')
     if target_id != 0:
       break
vk = vk_api.VkApi(token=key_vk)
#print(DiedTotalRF)
random_id=int(segodnyastr[20:])
message_vk='Обновлено : '+segodnyastr[:16]+ \
           '\nЗаразилось на сегодня = '+ str(CasesTotalRF)+ \
            '\nУмерло на сегодня = '+ str(DiedTotalRF)+ \
           ' чел \nСкорость заражения в день = '+str(Perday)[:6]+'%'\
           '\nПрогноз на завтра (экспонента) = '+ str(int(A1*numpy.exp(B1*(rownumber-1)))) +\
           '\nПрогноз на 1 мая (экспонента) = '+ str(int(A1*numpy.exp(B1*46)))+\
           '\nПрогноз на завтра (логистическая) = '+ str(int(f2(rownumber-1, A2, B2, C2,D2))) +\
           '\nПрогноз на 1 мая (логистическая) = '+ str(int(f2(46, A2, B2, C2,D2)))
#print(message_vk)
def write_msg(send_id, message, random_id):
    vk.method('messages.send', {'user_id': send_id, 'message': message, 'random_id': random_id})

def write_msg_chat(send_id, message, random_id):
    vk.method('messages.send', {'chat_id': send_id, 'message': message, 'random_id': random_id})
#print(otvet)
#write_msg_chat(chat_memy_tes, message_vk,random_id)
if target_id>100:
    write_msg(target_id, message_vk, random_id)
#elif otvet == 2 or 3:
if target_id<100:
     write_msg_chat(target_id, message_vk, random_id)

exit()
# чтобы проверять перед выходом цифры




